import dataset
from openpyxl import load_workbook
from sqlalchemy import Column, Integer, String, DateTime, Float, Boolean, LargeBinary
from sqlitedb import SQLiteDB
from sqlitedb.core import Base

class BundleIDs(Base):
    __tablename__ = "BundleIDs"
    ID = Column("ID", Integer, primary_key=True)
    App = Column("App", String)
    BundleID = Column("BundleID", String)
    Category = Column("Category", String)

def add_data():
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #data = list(row) + [""]
        #cols = ['App', "BundleID", "Category"]
        data = dict(App=row[0], BundleID=row[1], Category="")
        print(data)
        #db.insert(table_name, data)
        table.insert(data)

if __name__ == "__main__":
    excel_path = r"C:\Users\dwilliams\Code\Projects\jailbreak_tools\data\AppBundleIDs.xlsx"
    db_path = r"C:\Users\dwilliams\Code\Projects\jailbreak_tools\data\IconLibrary.db"
    db_path = "database.db"
    wb = load_workbook(excel_path)
    sheet = wb.active
    #db = SQLiteDB(db_path)
    db = dataset.connect(f"sqlite:///{db_path}")
    table_name = "BundleIDs"
    table = db[table_name]
    add_data()
