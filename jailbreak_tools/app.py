import os
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from pprint import pprint as pp

import dataset
from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets, uic

from . import CONFIG, DB, LOGGER, PATHS, __title__, __version__
from .models import *
from .ui import *


class App(QtWidgets.QApplication):
    sig_set_bundle_id = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__(sys.argv[1:])
        self.data = AppData()
        self.ui = IconLibrary(self, "iconlibrary")
        self.workbook = load_workbook(CONFIG.get("excel_file"))
        self.worksheet = self.workbook.active
        self.load()

    def load(self):
        self.connect_signals()
        self.connect_widgets()
        self.ui.show()

    def connect_signals(self):
        self.sig_set_bundle_id.connect(self.ui.input_bundle_id.setText)

    def connect_widgets(self):
        self.ui.input_app.editingFinished.connect(self.update_app_name)
        self.ui.btn_submit.clicked.connect(self.search_for_icon)

    def update_app_name(self):
        self.data.app_name = self.ui.input_app.text()
        print(self.data.app_name)
        self.update_bundle_id()

    def update_bundle_id(self):
        try:
            self.data.bundle_id = DB.get_bundle_id(self.data.app_name)
            self.sig_set_bundle_id.emit(self.data.bundle_id)
        except Exception as err:
            LOGGER.error(err)

    def get_data(self):
        return [row for row in self.worksheet.iter_rows(min_row=2, values_only=True)]

    def convert_excel_data_to_database(self):
        data = self.get_data()
        for row in data:
            icon_data = dict(App=row[0], BundleID=row[1], Category="")
            ##db.insert(table_name, data)
            self.table.insert(icon_data)
        LOGGER.info("BundleIDs Added to Database")

    def list_themes(self):
        theme_root = Path(CONFIG.get("theme_folder"))
        theme_dict = {}
        for theme in theme_root.iterdir():
            sub_dirs = len(list(theme.iterdir()))
            if sub_dirs > 2:
                for sub_theme in theme.iterdir():
                    if sub_theme.is_dir():
                        theme_dict[sub_theme.stem] = str(sub_theme)
            else:
                theme_dict[theme.stem] = str(theme)
        CONFIG.set("theme_library", theme_dict)
        CONFIG.save()

    def search_for_icon(self):
        for theme_name, theme_path in CONFIG.get("theme_library").items():
            theme = Path(os.path.join(theme_path, "IconBundles"))
            print(theme)
            for file in theme.iterdir():
                if self.data.bundle_id in file:
                    print(file)
            break
