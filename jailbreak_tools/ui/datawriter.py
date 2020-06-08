import os
import csv
from datetime import datetime
from platform import platform
from pathlib import Path, PurePath
from PyQt5 import QtCore, QtWidgets
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import PatternFill, colors
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import Workbook, load_workbook, formatting, styles
from openpyxl.chart import LineChart, ScatterChart, Series, Reference
from .. import CONFIG, LOGGER, paths, __title__


class DataWriter(QtCore.QObject):

    def __init__(self, app):
        super().__init__()
        self.app = app

    def create_data_dir(self):
        data_dir = CONFIG.get("data_dir")
        if not os.path.exists(data_dir):
            os.mkdir(data_dir)
            LOGGER.debug(f"{data_dir} Created")
        try:
            os.mkdir(self.app.data.file_dir)
            LOGGER.debug(f"{self.app.data.file_dir} Created")
        except Exception as err:
            LOGGER.error(err)

    def create_file(self, filename):
        with open(filename, "w+") as file:
            LOGGER.debug("{} Created".format(filename))
            file.close()

    def create_data_log(self):
        self.create_data_dir()
        self.create_file(self.app.data.file_path)
        sensor_header = ["", "", ""]
        sensor_info = ["", "", ""]
        data_header = ["Date", "Time", ""]
        spacer = [""]
        for station in self.app.data.active_stations:
            sensor_header += ["Comport", "Module Serial Number",
                              "Sensor Serial Number", ""]
            sensor_info += ["COM"+station.comport,
                            station.module_sn, station.sensor_sn, ""]
            data_header += ["ADC", "PPM", "Temp", ""]
        #interval_header = ["Temperature Interval", self.app.data.temp_interval]
        #duration_header = ["Test Duration", self.app.data.duration]
        #temp_points_header = ["Temperature Points"] + self.app.data.temp_points
        #self.write_data(self.app.data.filename, interval_header)
        #self.write_data(self.app.data.filename, duration_header)
        #self.write_data(self.app.data.filename, temp_points_header)
        self.write_data(self.app.data.filename, spacer)
        self.write_data(self.app.data.filename, sensor_info)
        self.write_data(self.app.data.filename, data_header)

    def create_sensor_log(self):
        sensor_header = ["Sensor Type",	"Module SN", "Sensor SN",
                         "Custom Baseline Table", "Gas Type ID",
                         "Date", "Initials", "Comments"]
        curr_date = datetime.now().strftime("%m-%d-%Y")
        for station in self.app.data.active_stations:
            sensor_info = [station.sensor_type, station.module_sn, station.sensor_sn,
                           '', station.gas_id, curr_date, self.app.data.tech_initials]
            title = station.module_sn
            if title == "":
                title = station.sensor_sn
            filename = "{}/{}.xlsx".format(self.app.data.file_dir, title)
            wb = Workbook()
            ws = wb.active
            ws.append(sensor_header)
            ws.append(sensor_info)
            ws.append([''])
            ws.append(["Temp", "ADC"])
            wb.save(filename)
            LOGGER.debug("{} Created".format(filename))

    def create_module_serial_numbers(self):
        self.create_data_dir()
        header = ["ModuleSN", "SensorSN"]
        filename = "{}/{}.xlsx".format(self.app.data.test_data_dir, self.app.data.board_module_sn_filename)
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for module_sn, sensor_sn in self.app.data.board_serial_numbers.items():
            ws.append([module_sn, sensor_sn])
        wb.save(filename)
        LOGGER.debug("{} Created".format(filename))

    def update_tempsweep_log(self, sensor_readings):
        curr_date = datetime.now().strftime("%m-%d-%Y")
        curr_time = datetime.now().strftime("%H:%M:%S %p")
        data = [curr_date, curr_time, ""]
        for reading in sensor_readings:
            data += reading
            data += ['']
        self.write_data(self.app.data.filename, data)

    def update_sensor_log(self):
        for station in self.app.data.active_stations:
            title = station.module_sn
            if title == "":
                title = station.sensor_sn
            filename = "{}/{}.xlsx".format(self.app.data.test_data_dir, title)
            wb = load_workbook(filename)
            ws = wb.active
            row = 5 + self.app.data.temp_interval_step
            ws.cell(row=row, column=1).value = int(station.temp)
            ws.cell(row=row, column=2).value = int(station.adc)
            wb.save(filename)
            LOGGER.debug("{} Updated".format(filename))

    def write_data(self, filename, data):
        with open(str(filename), 'a+', newline='') as csvfile:
            file = csv.writer(csvfile, delimiter=',')
            file.writerow(data)
            csvfile.close()

